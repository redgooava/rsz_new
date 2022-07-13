from django.http import HttpResponse
from openpyxl import Workbook
from .out import getDivisionList
from .models import Divisiontable, Outtable
from openpyxl.styles import Font, PatternFill, Border, Side, NamedStyle, Alignment
import datetime

styleTop = NamedStyle(name="topstyle")

styleTop.font = Font(name='Times New Roman',
                        size=14,
                        color='000000',
                        bold=False,
                        italic=False,
                        vertAlign=None,
                        underline='none',
                        strike=False)

styleTop.fill = PatternFill(fill_type='solid',
                               fgColor='A9A9A9')

styleTop.border = Border(left=Side(border_style='medium', color='000000'),
                            right=Side(border_style='medium', color='000000'),
                            top=Side(border_style='medium', color='000000'),
                            bottom=Side(border_style='medium', color='000000'),
                            diagonal=Side(border_style=None, color='000000'),
                            diagonal_direction=0,
                            outline=Side(border_style='medium', color='000000'),
                            vertical=Side(border_style='medium', color='000000'),
                            horizontal=Side(border_style='medium', color='000000'))

alignment_center = Alignment(horizontal='center')


def main(request, tableGeneral):
    today = datetime.datetime.now().date()
    filename = f"export/РСЗ_{today}.xlsx"
    if request.POST.get('docbutton') is not None:
        # response = HttpResponse(content_type='application/ms-excel')
        # response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb = Workbook()
        ws = wb.active
        create(ws)
        fillGeneral(tableGeneral, ws)
        fillOut(wb)
        wb.save(filename)
        # return response


def create(ws):
    ws.merge_cells('J1:T2')
    ws['J1'] = 'РАЗВЁРНУТАЯ СТРОЕВАЯ ЗАПИСКА за ______________ 20___ г.'

    for i in (10, 20):
        ws.cell(row=1, column=i).font = Font(name='Times New Roman', size=14, color='000000')
        ws.cell(row=2, column=i).font = Font(name='Times New Roman', size=14, color='000000')

    ws.merge_cells('A4:A5')
    ws['A4'] = 'Подразделение'
    ws['A4'].style = styleTop
    ws.merge_cells('B4:H4')
    ws['B4'] = 'По штату'
    ws['B4'].style = styleTop
    ws['C4'].style = styleTop
    ws['C4'].alignment = alignment_center
    ws.merge_cells('I4:O4')
    ws['I4'] = 'По списку'
    ws['I4'].style = styleTop
    ws.merge_cells('P4:V4')
    ws['P4'] = 'Налицо'
    ws['P4'].style = styleTop
    ws.merge_cells('W4:AB4')
    ws['W4'] = 'Отсутствуют'
    ws['W4'].style = styleTop

    ws['B5'] = 'Оф'
    ws['C5'] = 'Пр'
    ws['D5'] = 'К/с'
    ws['E5'] = 'С/с'
    ws['F5'] = 'К-ты'
    ws['G5'] = 'Сл'
    ws['H5'] = 'ВСЕГО'
    ws['I5'] = 'Оф'
    ws['J5'] = 'Пр'
    ws['K5'] = 'К/с'
    ws['L5'] = 'С/с'
    ws['M5'] = 'К-ты'
    ws['N5'] = 'Сл'
    ws['O5'] = 'ВСЕГО'
    ws['P5'] = 'Оф'
    ws['Q5'] = 'Пр'
    ws['R5'] = 'К/с'
    ws['S5'] = 'С/с'
    ws['T5'] = 'К-ты'
    ws['U5'] = 'Сл'
    ws['V5'] = 'ВСЕГО'
    ws['W5'] = 'Отпуск'
    ws['X5'] = 'Стац. лечение'
    ws['Y5'] = 'Амб. лечение'
    ws['Z5'] = 'Наряд'
    ws['AA5'] = 'Командировка'
    ws['AB5'] = 'ВСЕГО'

    for i in range(28):
        ws.cell(row=4, column=i + 1).style = styleTop
        ws.cell(row=5, column=i + 1).style = styleTop

    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['W'].width = 10
    ws.column_dimensions['X'].width = 17
    ws.column_dimensions['Y'].width = 17
    ws.column_dimensions['Z'].width = 10
    ws.column_dimensions['AA'].width = 17


def fillGeneral(table, ws):
    for row in table:
        ws.append(row)

    for i in range(len(table) + 5):
        for j in range(28):
            ws.cell(row=i+1, column=j+1).font = Font(name='Times New Roman', size=14, color='000000')


def fillOut(wb):
    divisionList = []
    preDivisionList = getDivisionList(Divisiontable.objects.all())
    wsOutAll = wb.create_sheet("Отсутствующие")
    for i in preDivisionList:
        divisionList.append(i['division'])
    for i in range(len(divisionList)):
        # wsOut = wb.create_sheet(f"{0}", divisionList[i][2])
        wsOut = wb.create_sheet(divisionList[i])
        divisionPageOut(divisionList[i],
                        Outtable.objects.filter(subdivision__exact=divisionList[i],
                                                datestart__lte=datetime.datetime.now(),
                                                dateend__gt=datetime.datetime.now(),
                                                daterealend__exact=None).values_list
                        ('name', 'rank', 'military_rank', 'subdivision', 'whyout',
                         'numberoforder', 'datestart', 'dateend', 'daterealend'),
                        wsOut, True)
    divisionPageOut(None, Outtable.objects.filter(datestart__lte=datetime.datetime.now(),
                                                  dateend__gt=datetime.datetime.now(),
                                                  daterealend__exact=None).values_list
                        ('name', 'rank', 'military_rank', 'subdivision', 'whyout',
                         'numberoforder', 'datestart', 'dateend', 'daterealend'),
                    wsOutAll, False)


def convertTable(queryset, ws):
    matrix = [] * len(queryset)
    for i in queryset:
        for j in i:
            matrix.append(j)

    for i in matrix:
        ws.append(i)


def divisionPageOut(division, tableOut, ws, isDivision):
    ws.merge_cells('A1:I1')
    ws['A1'] = 'Отсутствующие'

    if isDivision:
        ws.merge_cells('A2:I2')
        ws['A2'] = division

    ws['A3'] = 'ФИО'
    ws['B3'] = 'Категория'
    ws['C3'] = 'В/звание'
    ws['D3'] = 'Подразделение'
    ws['E3'] = 'Причина отсутствия'
    ws['F3'] = 'Номер приказа'
    ws['G3'] = 'Дата убытия'
    ws['H3'] = 'Дата прибытия (планируемая)'
    ws['I3'] = 'Дата прибытия (фактическая)'
    ws.merge_cells('I3:I3')  # костыль

    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 40
    ws.column_dimensions['I'].width = 40

    for i in range(9):
        ws.cell(row=1, column=i + 1).style = styleTop
        ws.cell(row=2, column=i + 1).style = styleTop
        ws.cell(row=3, column=i + 1).style = styleTop

    # for i in tableOut:
    #     i['datestart'] = dt.datetime.strptime(str(i['datestart']), '%Y-%m-%d').strftime('%d.%m.%Y')
    #     i['dateend'] = dt.datetime.strptime(str(i['datestart']), '%Y-%m-%d').strftime('%d.%m.%Y')
    #     i['daterealend'] = dt.datetime.strptime(str(i['datestart']), '%Y-%m-%d').strftime('%d.%m.%Y')

    fillGeneral(tableOut, ws)
    # convertTable(tableOut, ws)
